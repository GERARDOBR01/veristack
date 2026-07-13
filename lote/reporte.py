# -*- coding: utf-8 -*-
"""
lote/reporte.py — Reportería del modo lote: HTML visual auto-contenido
(miniaturas embebidas, GRAVES arriba, banner de parcialidad) y Excel de
datos (openpyxl; fallback a CSV — el reporte nunca tumba el lote).

Todo texto que viene del pipeline (razones del modelo incluidas) se escapa
con html.escape antes de entrar al HTML: el reporte se abre en navegador y
el contenido del modelo/knowledge es texto no confiable.

Uso:
  python -m lote.reporte autotest     (desde la raíz del repo)
"""

from __future__ import annotations

import base64
import html
import io
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

MINIATURA_MAX_PX = 240

_COLORES = {
    "GRAVE":       "#c0392b",
    "OBSERVACION": "#e67e22",
    "NO_CALIFICA": "#2980b9",
    "CUMPLE":      "#27ae60",
    "":            "#7f8c8d",   # no procesada / error
}
_ORDEN_REPORTE = {"GRAVE": 0, "OBSERVACION": 1, "NO_CALIFICA": 2, "": 3, "CUMPLE": 4}


# ──────────────────────────────────────────────────────────────
# MINIATURAS
# ──────────────────────────────────────────────────────────────

def _miniatura_data_uri(ruta: str) -> Optional[str]:
    """JPEG base64 redimensionado a MINIATURA_MAX_PX. None si no se puede leer
    (el reporte muestra un placeholder — nunca truena por una imagen)."""
    try:
        from PIL import Image
        with Image.open(ruta) as img:
            img = img.convert("RGB")
            img.thumbnail((MINIATURA_MAX_PX, MINIATURA_MAX_PX))
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=70)
        return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode("ascii")
    except Exception:
        return None


# ──────────────────────────────────────────────────────────────
# HTML
# ──────────────────────────────────────────────────────────────

def _e(texto) -> str:
    return html.escape(str(texto if texto is not None else ""))


def _bloque_foto(foto: dict, con_miniaturas: bool) -> str:
    veredicto = foto.get("veredicto_global", "") or ""
    color = _COLORES.get(veredicto, _COLORES[""])
    if foto.get("estado") == "no_procesada_por_cuota":
        etiqueta = "NO PROCESADA (cuota)"
    elif foto.get("estado") == "error":
        etiqueta = "ERROR DEL PIPELINE"
    else:
        etiqueta = veredicto or "—"

    mini = _miniatura_data_uri(foto.get("archivo", "")) if con_miniaturas else None
    img_html = (f'<img src="{mini}" alt="">' if mini
                else '<div class="ph">sin<br>miniatura</div>')

    filas = "".join(
        f"<tr><td>{_e(c['criterio'])}</td>"
        f"<td style='color:{_COLORES.get(c['veredicto'], '#000')};font-weight:600'>{_e(c['veredicto'])}</td>"
        f"<td>{_e(c['confianza'])}</td><td>{_e(c['fuente'])}</td><td>{_e(c['razon'])}</td></tr>"
        for c in foto.get("criterios_no_cumple", []))
    detalle = (f"<details><summary>{len(foto.get('criterios_no_cumple', []))} "
               f"criterio(s) con hallazgo — ver detalle</summary>"
               f"<table><tr><th>criterio</th><th>veredicto</th><th>confianza</th>"
               f"<th>fuente</th><th>razón</th></tr>{filas}</table></details>"
               if filas else "<p class='sin'>Sin hallazgos — todos los criterios CUMPLEN.</p>")

    parcial = ""
    if foto.get("evaluacion_parcial") and foto.get("estado") == "procesada":
        parcial = (f"<p class='parcial'>⚠️ EVALUACIÓN PARCIAL — "
                   f"{_e(foto.get('causa_parcial') or 'ver resumen')}</p>")

    conteo = (f"{foto.get('n_graves', 0)} graves · {foto.get('n_observaciones', 0)} obs. · "
              f"{foto.get('n_no_califica', 0)} no califica · {foto.get('n_cumple', 0)} cumplen")

    return f"""
<div class="foto" style="border-left:6px solid {color}">
  <div class="mini">{img_html}</div>
  <div class="info">
    <div class="cab"><span class="badge" style="background:{color}">{_e(etiqueta)}</span>
      <strong>{_e(foto.get('nombre', ''))}</strong>
      <span class="dur">{foto.get('duracion_s', 0)} s</span></div>
    {parcial}
    <p class="conteo">{_e(conteo)}</p>
    <p class="resumen">{_e(foto.get('resumen_ejecutivo', ''))}</p>
    {detalle}
  </div>
</div>"""


def generar_html(lote: dict, con_miniaturas: bool = True) -> str:
    """HTML auto-contenido (un solo archivo, sin recursos externos)."""
    meta, resumen = lote["meta"], lote["resumen"]
    pv = resumen["por_veredicto"]

    banner = ""
    if resumen.get("lote_parcial"):
        banner = (f"<div class='banner'>⚠️ EVALUACIÓN PARCIAL — este reporte NO es una "
                  f"corrida completa. Causa: {_e(resumen.get('causa_lote_parcial') or 'ver detalle por foto')}."
                  f" Fotos sin procesar por cuota: {resumen.get('fotos_no_procesadas_por_cuota', 0)}."
                  f" Fotos con error: {resumen.get('fotos_con_error', 0)}.</div>")

    tarjetas = "".join(
        f"<div class='card' style='border-top:4px solid {_COLORES[v]}'>"
        f"<div class='num'>{pv.get(v, 0)}</div><div>{v}</div></div>"
        for v in ("GRAVE", "OBSERVACION", "NO_CALIFICA", "CUMPLE"))
    pct = resumen.get("pct_cumplimiento")
    tarjetas += (f"<div class='card'><div class='num'>{pct if pct is not None else '—'}%</div>"
                 f"<div>cumplimiento</div></div>")

    fotos = sorted(lote["fotos"],
                   key=lambda f: (_ORDEN_REPORTE.get(f.get("veredicto_global", ""), 3),
                                  f.get("nombre", "")))
    bloques = "".join(_bloque_foto(f, con_miniaturas) for f in fotos)

    generado_local = datetime.now().strftime("%d %b %Y %H:%M")
    return f"""<!doctype html>
<html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Verificación de exhibiciones — {_e(meta.get('etapa_activa', ''))}</title>
<style>
 body{{font-family:'Segoe UI',system-ui,sans-serif;margin:0;background:#f4f5f7;color:#2c3e50}}
 header{{background:#2c3e50;color:#fff;padding:18px 28px}}
 header h1{{margin:0;font-size:1.3em}} header p{{margin:4px 0 0;opacity:.85;font-size:.9em}}
 .banner{{background:#c0392b;color:#fff;padding:14px 28px;font-weight:700}}
 .cards{{display:flex;gap:14px;padding:18px 28px;flex-wrap:wrap}}
 .card{{background:#fff;border-radius:8px;padding:12px 22px;text-align:center;
        box-shadow:0 1px 3px rgba(0,0,0,.12);min-width:110px}}
 .card .num{{font-size:1.7em;font-weight:700}}
 .foto{{display:flex;gap:16px;background:#fff;margin:12px 28px;border-radius:8px;
        padding:14px;box-shadow:0 1px 3px rgba(0,0,0,.12)}}
 .mini img{{max-width:{MINIATURA_MAX_PX}px;border-radius:6px;display:block}}
 .ph{{width:120px;height:90px;background:#dfe4ea;border-radius:6px;display:flex;
      align-items:center;justify-content:center;color:#7f8c8d;font-size:.8em;text-align:center}}
 .info{{flex:1;min-width:0}}
 .cab{{display:flex;align-items:center;gap:10px;flex-wrap:wrap}}
 .badge{{color:#fff;border-radius:4px;padding:2px 10px;font-weight:700;font-size:.85em}}
 .dur{{color:#95a5a6;font-size:.85em;margin-left:auto}}
 .parcial{{color:#c0392b;font-weight:700;margin:6px 0}}
 .conteo{{color:#7f8c8d;font-size:.9em;margin:6px 0}}
 .resumen{{font-size:.9em;margin:6px 0;overflow-wrap:break-word}}
 .sin{{color:#27ae60;font-weight:600}}
 details{{margin-top:6px}} summary{{cursor:pointer;color:#2980b9}}
 table{{border-collapse:collapse;width:100%;margin-top:8px;font-size:.85em}}
 th,td{{border:1px solid #dfe4ea;padding:6px 8px;text-align:left;vertical-align:top}}
 th{{background:#f4f5f7}}
 footer{{padding:14px 28px;color:#95a5a6;font-size:.8em}}
</style></head><body>
<header><h1>Verificación de exhibiciones — campaña {_e(meta.get('etapa_activa', ''))}</h1>
<p>{meta.get('fotos_procesadas', 0)} de {meta.get('fotos_totales', 0)} fotos procesadas ·
 tipo de foto: {_e(meta.get('tipo_foto_default', 'auto'))} · generado {_e(generado_local)}</p></header>
{banner}
<div class="cards">{tarjetas}</div>
{bloques}
<footer>Generado por Veristack / El Verificador · schema lote {_e(meta.get('schema_lote', ''))} ·
 orden: hallazgos graves primero. "NO_CALIFICA" = el sistema no pudo evaluar (distinto de detectar un problema).</footer>
</body></html>"""


# ──────────────────────────────────────────────────────────────
# EXCEL / CSV
# ──────────────────────────────────────────────────────────────

def _filas_resumen(lote: dict) -> list[dict]:
    return [{
        "archivo":            f.get("nombre", ""),
        "estado":             f.get("estado", ""),
        "veredicto":          f.get("veredicto_global", ""),
        "graves":             f.get("n_graves", 0),
        "observaciones":      f.get("n_observaciones", 0),
        "no_califica":        f.get("n_no_califica", 0),
        "cumplen":            f.get("n_cumple", 0),
        "evaluacion_parcial": "SI" if f.get("evaluacion_parcial") else "NO",
        "causa_parcial":      f.get("causa_parcial") or "",
        "duracion_s":         f.get("duracion_s", 0),
    } for f in lote["fotos"]]


def _filas_detalle(lote: dict) -> list[dict]:
    filas = []
    for f in lote["fotos"]:
        for c in f.get("criterios_no_cumple", []):
            filas.append({
                "archivo":   f.get("nombre", ""),
                "criterio":  c.get("criterio", ""),
                "veredicto": c.get("veredicto", ""),
                "confianza": c.get("confianza", ""),
                "fuente":    c.get("fuente", ""),
                "razon":     c.get("razon", ""),
            })
    return filas


def generar_excel(lote: dict) -> Optional[bytes]:
    """.xlsx con hojas Resumen y Detalle. None si falta pandas/openpyxl —
    el llamador cae a generar_csv_* (nunca fallar el lote por el reporte)."""
    try:
        import pandas as pd
        import openpyxl  # noqa: F401 — motor de to_excel; el import valida que está
    except ImportError:
        return None
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        pd.DataFrame(_filas_resumen(lote)).to_excel(xw, sheet_name="Resumen", index=False)
        detalle = _filas_detalle(lote)
        pd.DataFrame(detalle if detalle else [],
                     columns=["archivo", "criterio", "veredicto", "confianza", "fuente", "razon"]
                     ).to_excel(xw, sheet_name="Detalle", index=False)
    return buf.getvalue()


def _csv_bytes(filas: list[dict], columnas: list[str]) -> bytes:
    import csv
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=columnas, extrasaction="ignore")
    w.writeheader()
    w.writerows(filas)
    return out.getvalue().encode("utf-8-sig")   # BOM: Excel abre acentos bien


def generar_csv_resumen(lote: dict) -> bytes:
    return _csv_bytes(_filas_resumen(lote),
                      ["archivo", "estado", "veredicto", "graves", "observaciones",
                       "no_califica", "cumplen", "evaluacion_parcial", "causa_parcial", "duracion_s"])


def generar_csv_detalle(lote: dict) -> bytes:
    return _csv_bytes(_filas_detalle(lote),
                      ["archivo", "criterio", "veredicto", "confianza", "fuente", "razon"])


# ──────────────────────────────────────────────────────────────
# AUTOTEST — lote ficticio, cero pipeline, cero red
# ──────────────────────────────────────────────────────────────

def _lote_ficticio(parcial: bool) -> dict:
    return {
        "meta": {"schema_lote": "1.0", "generado": "2026-01-01T00:00:00+00:00",
                 "etapa_activa": "etapa_ficticia", "tipo_foto_default": "auto",
                 "fotos_totales": 2, "fotos_procesadas": 2 if not parcial else 1},
        "resumen": {"por_veredicto": {"GRAVE": 1, "OBSERVACION": 0, "NO_CALIFICA": 0,
                                      "CUMPLE": 0 if parcial else 1},
                    "pct_cumplimiento": 0.0 if parcial else 50.0,
                    "lote_parcial": parcial,
                    "causa_lote_parcial": "claves del modelo agotadas — FIXTURE" if parcial else None,
                    "fotos_parciales": 0, "fotos_no_procesadas_por_cuota": 1 if parcial else 0,
                    "fotos_con_error": 0},
        "fotos": [
            {"archivo": "no_existe_FIXTURE.jpg", "nombre": "no_existe_FIXTURE.jpg",
             "estado": "procesada", "veredicto_global": "GRAVE",
             "resumen_ejecutivo": "FIXTURE resumen <script>alert(1)</script>",
             "evaluacion_parcial": False, "causa_parcial": None, "duracion_s": 1.2,
             "criterios_no_cumple": [
                 {"criterio": "criterio_ficticio", "veredicto": "GRAVE", "confianza": "ALTO",
                  "fuente": "MANDATORY", "razon": "razón con <b>html</b> & ampersand"}],
             "n_graves": 1, "n_observaciones": 0, "n_no_califica": 0, "n_cumple": 3},
            ({"archivo": "b.jpg", "nombre": "b.jpg", "estado": "no_procesada_por_cuota",
              "veredicto_global": "", "resumen_ejecutivo": "cuota", "evaluacion_parcial": True,
              "causa_parcial": "cuota", "duracion_s": 0.0, "criterios_no_cumple": [],
              "n_graves": 0, "n_observaciones": 0, "n_no_califica": 0, "n_cumple": 0}
             if parcial else
             {"archivo": "b.jpg", "nombre": "b.jpg", "estado": "procesada",
              "veredicto_global": "CUMPLE", "resumen_ejecutivo": "todo bien",
              "evaluacion_parcial": False, "causa_parcial": None, "duracion_s": 0.8,
              "criterios_no_cumple": [], "n_graves": 0, "n_observaciones": 0,
              "n_no_califica": 0, "n_cumple": 4}),
        ],
    }


def autotest() -> int:
    fallas: list[str] = []

    def check(nombre: str, condicion: bool):
        print(f"  [{'PASS' if condicion else 'FAIL'}] {nombre}")
        if not condicion:
            fallas.append(nombre)

    # 1) HTML de lote completo: sin banner, GRAVE antes que CUMPLE, auto-contenido
    h = generar_html(_lote_ficticio(parcial=False))
    check("HTML lote completo: sin banner de parcialidad", "EVALUACIÓN PARCIAL — este reporte" not in h)
    check("HTML: GRAVE aparece antes que CUMPLE (orden de gravedad)",
          h.index("no_existe_FIXTURE.jpg") < h.index(">b.jpg<") if ">b.jpg<" in h
          else h.index("no_existe_FIXTURE.jpg") < h.index("b.jpg"))
    check("HTML: sin recursos externos (http solo en texto escapado)",
          'src="http' not in h and 'href="http' not in h and "@import" not in h)

    # 2) escape: el contenido del pipeline nunca entra como HTML vivo
    check("HTML: <script> del resumen queda escapado",
          "<script>alert(1)</script>" not in h and "&lt;script&gt;" in h)
    check("HTML: html en razones queda escapado", "<b>html</b>" not in h and "&lt;b&gt;" in h)

    # 3) imagen ilegible -> placeholder, nunca excepción
    check("miniatura de archivo inexistente -> None (placeholder)",
          _miniatura_data_uri("no_existe_FIXTURE.jpg") is None and 'class="ph"' in h)

    # 4) lote parcial: banner imposible de ignorar + foto no procesada visible
    hp = generar_html(_lote_ficticio(parcial=True), con_miniaturas=False)
    check("HTML parcial: banner con la causa",
          "EVALUACIÓN PARCIAL — este reporte NO es una corrida completa" in hp
          and "claves del modelo agotadas" in hp)
    check("HTML parcial: foto no procesada etiquetada", "NO PROCESADA (cuota)" in hp)

    # 5) Excel: 2 hojas relegibles (si openpyxl está; si no, se reporta y no falla)
    xlsx = generar_excel(_lote_ficticio(parcial=False))
    if xlsx is None:
        print("  [SKIP] Excel: pandas/openpyxl no instalados — fallback CSV cubre")
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        check("Excel: hojas Resumen y Detalle", wb.sheetnames == ["Resumen", "Detalle"])
        hoja = wb["Resumen"]
        check("Excel Resumen: 1 encabezado + 2 fotos", hoja.max_row == 3)
        check("Excel Detalle: fila del criterio ficticio",
              wb["Detalle"].cell(row=2, column=2).value == "criterio_ficticio")

    # 6) CSV fallback siempre disponible, BOM y filas correctas
    csv_r = generar_csv_resumen(_lote_ficticio(parcial=True))
    csv_d = generar_csv_detalle(_lote_ficticio(parcial=True))
    check("CSV resumen: BOM utf-8 + 2 filas de datos",
          csv_r.startswith(b"\xef\xbb\xbf") and csv_r.decode("utf-8-sig").strip().count("\n") == 2)
    check("CSV detalle: fila del criterio ficticio",
          "criterio_ficticio" in csv_d.decode("utf-8-sig"))

    # 7) el HTML es JSON-inerte (se genera desde dict serializable)
    check("lote ficticio JSON-serializable",
          json.loads(json.dumps(_lote_ficticio(True), ensure_ascii=False))["meta"]["fotos_totales"] == 2)

    print(f"\nAUTOTEST LOTE/REPORTE: {'PASS' if not fallas else 'FAIL'} ({len(fallas)} falla(s))")
    return len(fallas)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "autotest":
        sys.exit(1 if autotest() else 0)
    print("Uso: python -m lote.reporte autotest")
    sys.exit(2)
